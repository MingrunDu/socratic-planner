#!/usr/bin/env python3
"""Build a formatted literature-audit workbook from CSV inputs.

Inputs are optional except --evidence. Missing sheets are created with headers.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEETS = {
    "文献证据总表": [
        "模块",
        "文献标题",
        "作者年份",
        "期刊/机构",
        "文献类型",
        "研究对象/场景",
        "关键结果或参数",
        "与本文关系",
        "证据状态",
        "DOI/PMID/URL",
        "本地文件路径",
        "备注",
    ],
    "全文科学性审计": [
        "模块",
        "全文当前做法",
        "可比较/可引用文献",
        "科学性判断",
        "潜在风险/常识偏离",
        "优化建议",
        "优先级",
        "建议处理",
    ],
    "可新增或比较文献": [
        "推荐顺位",
        "文献标题",
        "发表年份",
        "作者年份",
        "期刊",
        "期刊等级/分区/IF",
        "为什么适合本文",
        "建议引用位置",
        "简述",
        "DOI/PMID/URL",
        "证据状态",
    ],
    "偏离常识风险清单": [
        "风险点",
        "为什么可能被质疑",
        "如何修正或解释",
        "严重性",
    ],
    "检索策略记录": [
        "检索时间",
        "数据库/来源",
        "检索式或关键词",
        "筛选规则",
        "命中数量",
        "纳入数量",
        "备注",
    ],
}


STATUS_FILL = {
    "verified": "D9EAD3",
    "local_verified": "D9EAD3",
    "partial": "FFF2CC",
    "reject": "F4CCCC",
    "high": "F4CCCC",
    "medium": "FFF2CC",
    "low": "D9EAD3",
}


def read_csv(path: str | None, headers: list[str]) -> pd.DataFrame:
    if not path:
        return pd.DataFrame(columns=headers)
    df = pd.read_csv(path)
    for col in headers:
        if col not in df.columns:
            df[col] = ""
    extra = [c for c in df.columns if c not in headers]
    return df[headers + extra]


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(["" if pd.isna(v) else v for v in row])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            value = str(cell.value).strip() if cell.value is not None else ""
            lower = value.lower()
            if lower in STATUS_FILL:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL[lower])
            if value.startswith("http://") or value.startswith("https://"):
                cell.hyperlink = value
                cell.style = "Hyperlink"

    if name == "可新增或比较文献":
        add_story_comments(ws)

    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in column_cells), default=0)
        header = str(ws.cell(row=1, column=idx).value or "")
        if header == "简述":
            ws.column_dimensions[get_column_letter(idx)].width = 36
        elif header == "简述详情":
            ws.column_dimensions[get_column_letter(idx)].hidden = True
        else:
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 55)


def add_story_comments(ws) -> None:
    headers = {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}
    summary_col = headers.get("简述")
    detail_col = headers.get("简述详情")
    if not summary_col:
        return

    for row_idx in range(2, ws.max_row + 1):
        summary_cell = ws.cell(row=row_idx, column=summary_col)
        detail = ""
        if detail_col:
            detail = str(ws.cell(row=row_idx, column=detail_col).value or "").strip()
        visible = str(summary_cell.value or "").strip()
        full_text = detail or visible
        if not full_text:
            continue

        if len(visible) > 120:
            summary_cell.value = visible[:117].rstrip() + "..."
        elif detail and not visible:
            first_line = detail.splitlines()[0].strip()
            summary_cell.value = first_line[:117].rstrip() + ("..." if len(first_line) > 120 else "")

        summary_cell.comment = Comment(full_text, "socratic-paper-finder")
        summary_cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 36


def build(args: argparse.Namespace) -> None:
    mapping = {
        "文献证据总表": args.evidence,
        "全文科学性审计": args.audit,
        "可新增或比较文献": args.recommended,
        "偏离常识风险清单": args.risks,
        "检索策略记录": args.search_log,
    }
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name, headers in SHEETS.items():
        write_sheet(wb, name, read_csv(mapping.get(name), headers))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--evidence", required=True, help="CSV for 文献证据总表")
    parser.add_argument("--audit", help="CSV for 全文科学性审计")
    parser.add_argument("--recommended", help="CSV for 可新增或比较文献")
    parser.add_argument("--risks", help="CSV for 偏离常识风险清单")
    parser.add_argument("--search-log", help="CSV for 检索策略记录")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    return parser.parse_args(argv)


if __name__ == "__main__":
    build(parse_args())
